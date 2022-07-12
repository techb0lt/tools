import streamlit as st
import pandas as pd
import numpy as np
from pandas import ExcelWriter
from podb_functions import read_upload, to_sql, run_sql, run_sql_file, read_sql, generic_aggrid
import re
from st_aggrid import AgGrid, GridOptionsBuilder

# Set wide display
st.set_page_config(
    layout="wide",
)

################## Objective ########################################
# Select Excel 1                                                    #
# Select Excel 2                                                    #
#                                                                   #
# if Excel 1 exists:                                                #
#    Obtain Sheets in Excel 1                                       #
#    Show Sheets in dropdown                                        #
# if Excel 2 exists:                                                #
#    Obtain Sheet in Excel 2                                        #
#    Show Sheets in dropdown                                        #
#                                                                   #
# User must select a Sheet from each of the above dropdowns         #
#                                                                   #
# Using Sheetname, read data from that sheet into a dataframe.      #
#                                                                   #
# Provide dropdown to show column names from each dataset           #
#                                                                   #
# User must select the columns to compare                           #
#                                                                   #
# Comparison should then show for the specified column Venn Diagram #
# and data with exclusive entries in each dataset.                  #
#####################################################################


st.markdown("""
# Guidance Notes

**Set Comparison Tool** : The idea behind this tool is to simply help the user to identify from two given sets of values the ones that are common, and ones that are exclusive to each set. The set values are provided by means of selecting a column each from the two spreadsheets that user is able to upload using the upload button.


> :bulb: If the comparison is required between two columns of same spreadsheet, then the same spreadsheet can be uploaded twice and then from the dropdown for columns user can select the columns to compare.
""")

def get_sheetnames(upload_file):
    from openpyxl import load_workbook
    wb = load_workbook(upload_file, read_only=True, keep_links=False)
    return wb.sheetnames

file_upload_container = st.container()
sheet_selection_container = st.container()
column_selection_container = st.container()
answer_container = st.container()
answer_tables = st.container()

with file_upload_container:
    f_col1, f_col2 = st.columns(2)
    f_button = st.empty()
with sheet_selection_container:
    s_col1, s_col2 = st.columns(2)
    s_button = st.empty()
with column_selection_container:
    c_col1, c_col2 = st.columns(2)
    c_button = st.empty()
with answer_container:
    download_button_placeholder = st.empty()
    venn_placeholder, a_col1 = st.columns(2)
with answer_tables:
    a_t_col1,a_t_col2 = st.columns(2)

with f_col1:
    f1_uploader = st.file_uploader(
        "Choose file 1",
        help="Upload the first file in .xls or .xlsx format",
        type=['xlsx','xls'])
    error_placeholder_f1 = st.empty()
with f_col2:
    f2_uploader = st.file_uploader(
        "Choose file 2",
        help="Upload the second file in .xls or .xlsx format",
        type=['xlsx','xls'])
    error_placeholder_f2 = st.empty()

if f1_uploader is not None and f2_uploader is not None:
    # Read sheet names for both files
    f1_sheets = get_sheetnames(f1_uploader)
    f1_sheets_mod = ("Select",*f1_sheets)
    f2_sheets = get_sheetnames(f2_uploader)
    f2_sheets_mod = ("Select",*f2_sheets)
    with sheet_selection_container:
        with s_col1:
            sheet1 = st.selectbox("File 1 Sheets",f1_sheets_mod, help = "Select a sheet from this list")
        with s_col2:
            sheet2 = st.selectbox("File 2 Sheets",f2_sheets_mod, help = "Select a sheet from this list")
        if sheet1 == "Select" or sheet2 == "Select":
            st.warning('Sheets from both files must be selected for comparison.')
            st.stop()
        else:
            with s_col1:
                #st.write(sheet1)
                sheet1_data = read_upload(f1_uploader,sheet1)

            with s_col2:
                #st.write(sheet2)
                sheet2_data = read_upload(f2_uploader,sheet2)

        with column_selection_container:
            if isinstance(sheet1_data,pd.DataFrame) and isinstance(sheet2_data,pd.DataFrame):
                with c_col1:
                    sheet1_cols_mod = ("Select",*sheet1_data.columns.values.tolist())
                    set1_name = st.selectbox('Set 1',sheet1_cols_mod,help = "Select Column for set 1 of the comparison")
                with c_col2:
                    sheet2_cols_mod = ("Select",*sheet2_data.columns.values.tolist())
                    set2_name = st.selectbox('Set 2',sheet2_cols_mod,help = "Select Column for set 2 of the comparison")
            if set1_name == "Select" or set2_name == "Select":
                st.warning('Sets from both files must be selected for comparison.')
                st.stop()
            else:
                with c_col1:
                    set1 = set(sheet1_data[set1_name])
                    set2 = set(sheet2_data[set2_name])
                    with st.expander(set1_name, expanded=False):
                        st.write(set1)
                with c_col2:
                    #st.write(set2)
                    with st.expander(set2_name, expanded=False):
                        st.write(set2)
                with answer_container:
                    with venn_placeholder:
                        import matplotlib.pyplot as plt
                        from matplotlib_venn import venn2, venn2_circles
                        fig, ax = plt.subplots()
                        ax = venn2([set1, set2], (sheet1 + '.' + set1_name+'(Total '+str(len(set1))+')', sheet2 + '.' + set2_name + ' (Total '+str(len(set2))+')'))
                        ax.get_patch_by_id('10').set_color('red')
                        ax.get_patch_by_id('01').set_color('blue')
                        if ax.get_patch_by_id('11') is not None:
                            ax.get_patch_by_id('11').set_color('green')
                        plt.title("Venn Diagram")
                        st.pyplot(fig)
                    with a_col1:
                        st.markdown("""
                        ||{5}|{6}|
                        |---|---|---|
                        |Total|{3}|{4}|
                        |Exclusive (Only in)|{2}|{1}|
                        |Common to Both|{0}|{0}|

                        """.format(len(set2.intersection(set1)),
                                   len(set2.difference(set1)),
                                   len(set1.difference(set2)),
                                   len(set1),
                                   len(set2),
                                   sheet1 +'.'+ set1_name,
                                   sheet2 +'.'+ set2_name
                                  )
                        )
                    with a_t_col1:
                        st.markdown("The {0} entries exclusive to {1} are:".format(
                        len(set1.difference(set2)),sheet1 +'.'+ set1_name
                        ))
                        data = sheet1_data.loc[sheet1_data[set1_name].isin(set1.difference(set2))]
                        generic_aggrid(data)
                        
                    with a_t_col2:
                        st.markdown("The {0} entries exclusive to {1} are:".format(
                        len(set2.difference(set1)),sheet2 +'.'+ set2_name
                        ))
                        generic_aggrid(data = sheet2_data.loc[sheet2_data[set2_name].isin(set2.difference(set1))])
                        #st.write(set1.difference(set2))
elif f1_uploader is None and f2_uploader is None:
    #don't show anything.
    print("App has just started.")
elif f1_uploader is None:
    with error_placeholder_f1:
        st.error("Please upload file 1.")
else:
    with error_placeholder_f2:
        st.error("Please upload file 2.")
